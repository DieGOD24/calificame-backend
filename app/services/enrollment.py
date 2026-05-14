import csv
import io
import re
import unicodedata

from fastapi import UploadFile
from loguru import logger
from sqlalchemy.orm import Session

from app.models.user import User

NAME_ALIASES = {
    "nombre",
    "nombres",
    "name",
    "estudiante",
    "estudiantes",
    "student",
    "alumno",
    "alumnos",
    "student_name",
    "apellidos y nombres",
    "nombre completo",
    "nombres y apellidos",
}

ID_ALIASES = {
    "codigo",
    "código",
    "code",
    "identifier",
    "id",
    "matricula",
    "matrícula",
    "student_identifier",
    "cedula",
    "cédula",
    "cc",
    "dni",
    "documento",
    "documento de identidad",
    "doc",
    "identificacion",
    "identificación",
    "numero de documento",
    "número de documento",
    "no. documento",
    "nro documento",
    "nro. documento",
}

EMAIL_ALIASES = {
    "email",
    "e-mail",
    "correo",
    "correo electronico",
    "correo electrónico",
    "mail",
    "student_email",
}

EMAIL_RE = re.compile(r"[\w\.\-\+]+@[\w\.\-]+\.\w+")


async def parse_student_file(file: UploadFile) -> list[dict]:
    """Parse a CSV or XLSX file and return a list of student records.

    Accepts varied layouts — metadata rows before the header are tolerated.
    Expected columns (or Spanish equivalents): name, identifier/document, email.
    Returns `[]` if nothing could be extracted; caller may fall back to AI.
    """
    content = await file.read()
    filename = (file.filename or "").lower()

    if filename.endswith(".xlsx"):
        return _parse_xlsx(content)
    return _parse_csv(content)


def _normalize(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    # Strip accents for matching (keeps original for display elsewhere)
    return "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")


def _find_header_row(rows: list[list]) -> int | None:
    """Scan the first ~30 rows for one that looks like a header row.

    A header row is one with at least name+identifier alias matches.
    """
    limit = min(len(rows), 30)
    for i in range(limit):
        row = rows[i]
        norm = [_normalize(c) for c in row]
        has_name = any(cell in NAME_ALIASES for cell in norm)
        has_id = any(cell in ID_ALIASES for cell in norm)
        if has_name and has_id:
            return i
    return None


def _build_header_map(headers: list[str]) -> dict[str, int]:
    """Map standardized keys ('name', 'identifier', 'email') to column indexes."""
    mapping: dict[str, int] = {}
    for idx, header in enumerate(headers):
        norm = _normalize(header)
        if "name" not in mapping and norm in NAME_ALIASES:
            mapping["name"] = idx
        if "identifier" not in mapping and norm in ID_ALIASES:
            mapping["identifier"] = idx
        if "email" not in mapping and norm in EMAIL_ALIASES:
            mapping["email"] = idx
    return mapping


def _extract_email(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    match = EMAIL_RE.search(text)
    return match.group(0).lower() if match else None


def _parse_rows(rows: list[list]) -> list[dict]:
    """Shared logic: find header row, map columns, extract records."""
    if not rows:
        return []

    header_idx = _find_header_row(rows)
    if header_idx is None:
        return []

    headers = [str(h) if h is not None else "" for h in rows[header_idx]]
    header_map = _build_header_map(headers)
    if "name" not in header_map or "identifier" not in header_map:
        return []

    results: list[dict] = []
    for row in rows[header_idx + 1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        name = str(row[header_map["name"]] or "").strip() if header_map["name"] < len(row) else ""
        identifier_raw = row[header_map["identifier"]] if header_map["identifier"] < len(row) else None
        identifier = str(identifier_raw).strip() if identifier_raw is not None else ""

        # Excel sometimes stores numeric IDs as floats ("1116434602.0") — trim trailing .0
        if identifier.endswith(".0") and identifier[:-2].isdigit():
            identifier = identifier[:-2]

        email: str | None = None
        if "email" in header_map and header_map["email"] < len(row):
            email = _extract_email(row[header_map["email"]])
        if email is None:
            # Fallback: scan the whole row for an email-looking token
            for cell in row:
                email = _extract_email(cell)
                if email:
                    break

        if name and identifier:
            results.append({
                "student_name": name,
                "student_identifier": identifier,
                "student_email": email,
            })

    return results


def _parse_csv(content: bytes) -> list[dict]:
    """Parse CSV content into student records."""
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = content.decode("latin-1", errors="replace")

    if not text.strip():
        raise ValueError("El archivo CSV está vacío")

    reader = csv.reader(io.StringIO(text))
    rows = [row for row in reader]
    return _parse_rows(rows)


def _parse_xlsx(content: bytes) -> list[dict]:
    """Parse XLSX content into student records."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ValueError("openpyxl no está instalado para procesar archivos Excel") from e

    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"No se pudo abrir el archivo Excel: {e}") from e

    all_records: list[dict] = []
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        records = _parse_rows(rows)
        if records:
            all_records.extend(records)

    wb.close()
    return all_records


def flatten_to_text(content: bytes, filename: str) -> str:
    """Flatten any supported file to a plain-text table representation for AI parsing."""
    lower = filename.lower()
    if lower.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            return ""
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception:
            return ""
        lines: list[str] = []
        for ws in wb.worksheets:
            lines.append(f"--- Sheet: {ws.title} ---")
            for row in ws.iter_rows(values_only=True):
                cells = ["" if c is None else str(c) for c in row]
                lines.append("\t".join(cells))
        wb.close()
        return "\n".join(lines)

    # CSV / plain text
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("latin-1", errors="replace")


def auto_link_users(db: Session, enrollments: list[dict]) -> list[dict]:
    """Try to match student emails with existing User accounts."""
    emails = [e["student_email"] for e in enrollments if e.get("student_email")]
    if not emails:
        return enrollments

    users = db.query(User).filter(User.email.in_(emails)).all()
    email_to_user = {u.email: u.id for u in users}

    for enrollment in enrollments:
        email = enrollment.get("student_email")
        if email and email in email_to_user:
            enrollment["user_id"] = email_to_user[email]
            logger.info(f"Auto-linked student {enrollment['student_identifier']} to user {email}")

    return enrollments
