import csv
import io

from sqlalchemy.orm import Session

from app.models.clase import Class, ClassEnrollment, ClassProject
from app.models.student_exam import StudentExam
from app.schemas.clase import (
    GradebookCell,
    GradebookResponse,
    GradebookRow,
    StudentProgressResponse,
)

PASS_THRESHOLD = 60.0


def build_gradebook(db: Session, clase: Class) -> GradebookResponse:
    """Build a gradebook for a class: students (rows) x projects (columns)."""
    enrollments = (
        db.query(ClassEnrollment)
        .filter(ClassEnrollment.class_id == clase.id)
        .order_by(ClassEnrollment.student_name)
        .all()
    )

    class_projects = (
        db.query(ClassProject).filter(ClassProject.class_id == clase.id).order_by(ClassProject.display_order).all()
    )

    # Build column names
    columns: list[str] = []
    project_ids: list[str] = []
    project_names: dict[str, str] = {}
    for cp in class_projects:
        proj = cp.project
        name = proj.name if proj else f"Proyecto {cp.project_id[:8]}"
        columns.append(name)
        project_ids.append(cp.project_id)
        project_names[cp.project_id] = name

    # Fetch all student exams for these projects. There's a UNIQUE constraint
    # on (project_id, student_identifier), so at most one exam exists per
    # (student, project) — the order_by/preference logic that used to live
    # here was a no-op, but it could mask a status=error row that still had
    # a stale grade_percentage from an earlier successful grading run (the
    # Jorge Luis incident — see QA report S2.2). The fix: read `status`
    # explicitly downstream and never expose grade_percentage when status is
    # not "graded".
    exam_lookup: dict[tuple[str, str], StudentExam] = {}
    if project_ids:
        exams = (
            db.query(StudentExam)
            .filter(
                StudentExam.project_id.in_(project_ids),
                StudentExam.student_identifier.isnot(None),
            )
            .all()
        )
        for exam in exams:
            key = (exam.student_identifier.strip().lower(), exam.project_id)
            exam_lookup[key] = exam

    # Build rows
    rows: list[GradebookRow] = []
    for enrollment in enrollments:
        student_id_norm = enrollment.student_identifier.strip().lower()
        cells: list[GradebookCell] = []
        scores: list[float] = []

        for pid in project_ids:
            exam = exam_lookup.get((student_id_norm, pid))
            if exam and exam.status == "graded" and exam.grade_percentage is not None:
                cells.append(
                    GradebookCell(
                        project_id=pid,
                        project_name=project_names[pid],
                        score=exam.total_score,
                        max_score=exam.max_score,
                        percentage=exam.grade_percentage,
                        status="graded",
                    )
                )
                scores.append(exam.grade_percentage)
            else:
                # Surface the underlying status so the UI can render an
                # "error" / "processing" / "uploaded" badge instead of an
                # empty cell that looks identical to "no exam at all".
                cells.append(
                    GradebookCell(
                        project_id=pid,
                        project_name=project_names[pid],
                        status=exam.status if exam else None,
                        error_message=(
                            exam.error_message if exam and exam.status == "error" else None
                        ),
                    )
                )

        avg = round(sum(scores) / len(scores), 1) if scores else None
        if avg is not None:
            pass_status = "passing" if avg >= PASS_THRESHOLD else "failing"
        else:
            pass_status = "pending"

        rows.append(
            GradebookRow(
                student_name=enrollment.student_name,
                student_identifier=enrollment.student_identifier,
                projects=cells,
                average=avg,
                pass_status=pass_status,
            )
        )

    return GradebookResponse(
        class_id=clase.id,
        class_name=clase.name,
        semester=clase.semester,
        columns=columns,
        rows=rows,
    )


def get_student_progress(db: Session, clase: Class, enrollment: ClassEnrollment) -> StudentProgressResponse:
    """Get a single student's progress across all class projects."""
    gradebook = build_gradebook(db, clase)
    student_row = next(
        (r for r in gradebook.rows if r.student_identifier == enrollment.student_identifier),
        None,
    )
    if student_row:
        return StudentProgressResponse(
            student_name=student_row.student_name,
            student_identifier=student_row.student_identifier,
            class_name=clase.name,
            semester=clase.semester,
            projects=student_row.projects,
            average=student_row.average,
        )
    return StudentProgressResponse(
        student_name=enrollment.student_name,
        student_identifier=enrollment.student_identifier,
        class_name=clase.name,
        semester=clase.semester,
        projects=[],
        average=None,
    )


def export_gradebook_csv(gradebook: GradebookResponse) -> bytes:
    """Export gradebook as CSV bytes."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Header row
    header = ["Estudiante", "Codigo", *gradebook.columns, "Promedio", "Estado"]
    writer.writerow(header)

    # Data rows
    for row in gradebook.rows:
        data = [row.student_name, row.student_identifier]
        for cell in row.projects:
            data.append(f"{cell.percentage:.1f}%" if cell.percentage is not None else "-")
        data.append(f"{row.average:.1f}%" if row.average is not None else "-")
        status_map = {"passing": "Aprobado", "failing": "Reprobado", "pending": "Pendiente"}
        data.append(status_map.get(row.pass_status, row.pass_status))
        writer.writerow(data)

    return output.getvalue().encode("utf-8-sig")


def export_gradebook_xlsx(gradebook: GradebookResponse) -> bytes:
    """Export gradebook as XLSX bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = f"{gradebook.class_name} - {gradebook.semester}"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    green_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")

    # Header
    headers = ["Estudiante", "Codigo", *gradebook.columns, "Promedio", "Estado"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data
    for row_idx, row in enumerate(gradebook.rows, 2):
        ws.cell(row=row_idx, column=1, value=row.student_name)
        ws.cell(row=row_idx, column=2, value=row.student_identifier)

        for col_idx, cell_data in enumerate(row.projects, 3):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell_data.percentage is not None:
                cell.value = round(cell_data.percentage, 1)
                cell.number_format = "0.0"
                if cell_data.percentage >= 80:
                    cell.fill = green_fill
                elif cell_data.percentage >= 60:
                    cell.fill = yellow_fill
                else:
                    cell.fill = red_fill
            else:
                cell.value = "-"
            cell.alignment = Alignment(horizontal="center")

        avg_col = len(row.projects) + 3
        avg_cell = ws.cell(row=row_idx, column=avg_col)
        if row.average is not None:
            avg_cell.value = round(row.average, 1)
            avg_cell.number_format = "0.0"
        else:
            avg_cell.value = "-"
        avg_cell.alignment = Alignment(horizontal="center")

        status_map = {"passing": "Aprobado", "failing": "Reprobado", "pending": "Pendiente"}
        ws.cell(row=row_idx, column=avg_col + 1, value=status_map.get(row.pass_status, row.pass_status))

    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
